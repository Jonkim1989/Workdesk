# 리포트에 분석표, 검색순위 제거 이후 정렬과 해더만 꾸미기

#----------------------------------------------------------
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, colors, Font, Alignment
import openpyxl
import os
from os import walk

path_1 = r"C:\Users\jonkim\Desktop\리포트 마무리\1. 작업 대기\pd작업"

f = []  

# 작업 폴더에 있는 파일의 filenames 만 리스트에 저장   
for (dirpath, dirnames, filenames) in walk(path_1):
    f.extend(filenames)
    break # 이 break가 꼭 필요할까?
 
for full_name in f:
    wb = openpyxl.load_workbook(path_1 + "\\" + full_name)
    ws = wb.active

    # 1차 가공 - 셀 정렬, 크기 조절, 폰트 변경

    for row in ws.rows:
        for cell in row:
            # 셀 폰트 변경
            cell.font = Font(name="Calibri", size=11)
            
            # D.A column 중앙 정렬
            if cell.column == 3:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # column 크기 조절
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 130



    # 해더 색상 입히기
    for row in ws["A1:D1"]:
        for cell in row:
            cell.font = Font(name="Calibri", size=11, color=colors.WHITE, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="bottom")
            cell.fill = PatternFill(start_color="366092", fill_type="solid")

    
    # 저장하기
    wb.save(path_1 + "\\" + full_name)
