from asyncio import futures
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, colors, Font, Alignment
import openpyxl
import os
from os import walk



# opening the source excel file
wb1= openpyxl.load_workbook("Report_test.xlsx")
ws1 = wb1.worksheets[0]

# opening the destination excel file 
wb2= openpyxl.load_workbook("리포트 마무리 템플릿.xlsx")
ws2 = wb2.active

# calculate total number of rows and 
# columns in source excel file
mr = ws1.max_row
mc = ws1.max_column
  
# copying the cell values from source 
# excel file to destination excel file
for i in range (1, mr + 1):
    ws2.row_dimensions[i+10].height = 16.5
    for j in range (1, mc + 1):
        # reading cell value from source excel file
        c = ws1.cell(row = i, column = j)
  
        # writing the read value to destination excel file
        ws2.cell(row = i + 10, column = j).value = c.value
        ws2.cell(row = i + 10, column = j).font = Font(name="나눔고딕", size=10)
        if c.column == 3:
            c.alignment = Alignment(horizontal="center")


# A열에서 특정 값을 가진 첫번 째 셀 row 좌표 찾기
label_list = ["Web 2.0 Blogs", "Social Bookmarking", "Authority Links", "Edu", "URL Shortener"]

for label in label_list:
    temp_x = ()
    for row in ws2.iter_rows():
        temp_i = row[0].value.split(" (")[0]
        if temp_i == label:
            temp_x = row[0].row
            break

    ws2.insert_rows(temp_x)  # temp_x위에 빈줄 추가

    for i in ("A", "B", "C", "D"):
        ws2[f"{i}{temp_x}"].font = Font(name="나눔고딕", size=8.5, color=colors.WHITE, bold=True)
        ws2[f"{i}{temp_x}"].alignment = Alignment(horizontal="center", vertical="center")
        ws2[f"{i}{temp_x}"].fill = PatternFill(start_color="366092", fill_type="solid")

    data = {1: label, 2: "Site URL", 3: "D.A.", 4: "Submitted URL"}
    for col, value in data.items():
        ws2.cell(row=temp_x, column=col, value=value)

# 마지막 줄에 푸터 삽입
footer_x = ws2.max_row+1
ws2.merge_cells(f"A{footer_x}:D{footer_x}")  # 병합 하기
ws2[f"A{footer_x}"] = "마케스터즈 | Marketsters | https://marketsters.com/"
ws2[f"A{footer_x}"].font = Font(name="나눔고딕", size=8.5, color=colors.WHITE, bold=True)
ws2[f"A{footer_x}"].alignment = Alignment(vertical="center")
ws2[f"A{footer_x}"].fill = PatternFill(start_color="366092", fill_type="solid")
    
 
# saving the destination excel file
wb2.save("리포트 마무리 템플릿2.xlsx")


# 기존 리포트 템플릿을 불러오고 1차 정리된 백링크를 for 문으로 한 셀씩 채워 넣는 방식. 
# 이게 해드를 매번 만드는것과 비교해서 빠른지 테스트 필요하고 
# task 가 바뀔때마다 라벨을 만들어줘야 한다.fd





