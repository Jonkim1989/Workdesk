from openpyxl.drawing.image import Image
from openpyxl import load_workbook  # 대문자로 적어야 모듈이 활성화 됨
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils.cell import coordinate_from_string  # 작업하고자 하는 셀의 위치를 표시하게 해줌
# xy = coordinate_from_string(cell.coordinate)

wb = load_workbook("D12-28 casinopan M-G-4-1.xlsx")
ws = wb.active

# cell 데이터 불러오기
# for x in range(1, ws.max_row + 1):
#     for y in range(1, ws.max_column + 1):
#         # 반복해서 출력할때 끝에 end= " " 가 없으면 다음줄로 출력됨
#         print(ws.cell(row=x, column=y).value, end=" ")
#     print()

# cell 데이터 찾기
# for row in ws.iter_rows(min_row=3):
#     if row[1].value == "wpsuo.com":
#         print(row[1].value, "이게 출력된다.")

# 특정 위치에 빈 row 와 column을 추가한다.
ws.insert_rows(8)  # 8번째 행에 빈줄 추가
ws.insert_rows(8, 5)  # 8번 행 부터 빈줄 5개 추가

ws.insert_cols(8)  # 8번째 열에 빈줄 추가
ws.insert_cols(8, 5)  # 8번 열 부터 빈줄 5개 추가

# 특정 위치에 빈 row 와 column을 삭제한다.
ws.delete_rows(8)  # 8번째 행 삭제 후 자동 당김
ws.delete_rows(8, 5)  # 8번 행 부터 줄 5개 삭제 후 자동 당김

ws.delete_cols(8)  # 8번째 열 삭제 후 자동 당김
ws.delete_cols(8, 5)  # 8번 열 부터 줄 5개 삭제 후 자동 당김

# 기존에 있는 row 와 column을 이동시킨다. (내용을 덮어 씌움)
ws.move_range("B11:C11", rows=0, cols=1)


# 행 열 크기 바꾸기
ws.column_dimensions["A"].width = 5  # A열의 넓이를 5만큼 변경
ws.row_dimensions[1].width = 5  # 1번 행의 넓이를 5만큼 변경

# 스타일 적용
ws["A1"].font = Font(color="FF0000", italic=True,
                     bold=True)  # 셀의 글자 색, 이탈릭체, 두껍게 적용
ws["B1"].font = Font(color="FF0000", name="Arial",
                     strike=True)  # 셀의 글자 색, 폰트 변경, 중간줄 적용
# 셀의 글자 색, 사이즈 변경, 및줄 적용
ws["C1"].font = Font(color="FF0000", size=20, underline="single")


# 테두리 적용
thin_border = Border(left=Side(style="thin"), right=Side(
    style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
ws["A1"].border = thin_border

# 셀 중앙 정렬  (center, left, right, top, bottom등을 설정 할 수 있다.) for 문으로 각 셀을 하나씩 정렬하도록 해야 하나보다.
ws["C1"].alignment = Alignment(horizontal="center", vertical="center")

# 틀 고정
ws.freeze_panes = "B2"  # B2를 기준으로 틀을 행과 열을 고정시킨다.


# 셀 병합, 병합 해제
ws.merge_cells("B2:B3")  # 병합 하기
ws.unmerge_cells("B2:B3")  # 병합 해지 하기

# 이미지 가져오기

img = Image("img.png")
ws.add_image(img, "C3")  # 이미지를 C3 위치에 삽입


wb.save("new casino.xlsx")
