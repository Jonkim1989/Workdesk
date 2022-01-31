from openpyxl import Workbook # 대문자로 적어야 모듈이 활성화 됨

wb = Workbook() # 새 워크북 생성
# ws = wb.active # 현재 활성화된 sheet 가져옴
# ws.title = "test_sample" # sheet 의 이름 변경

ws = wb.create_sheet() # 기본 이름으로 시트 생성
ws1 = wb.create_sheet("Name") # 새로운 시트를 name1이란 이름으로 생성

TS = wb["Name"] # 특정 시트를 이름으로 불러오기

TS["A1"] = "글쓰기" # 셀 안에 글쓰기

print(TS["A1"].value) # 딕셔너리 형식으로 값을 출력하기
print(TS.cell(column=1, row=1).value) # 셀의 위치를 지정해서 값을 출력하기








wb.save("test_1.xlsx") # 엑셀파일 저장
wb.close() # 마지막으로 오크북 닫기


