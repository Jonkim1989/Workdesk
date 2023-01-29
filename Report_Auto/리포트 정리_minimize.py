# 2023.01.24 리포트 정리 코드 
# Task ID, Submitted URL 제외하고 정리하기 
# task ID = Tier # 로 변경
# Submitted URL = Backlink List URL 로 변경
# Tier # 행에 있는 모든 작업 이름들 제거 티어만 남기기

#------------------------------------------------------------------------


from openpyxl.styles import PatternFill, colors, Font, Alignment
import openpyxl
from os import walk
import pandas as pd

path_1 = r"C:\Users\jonkim\Desktop\리포트 마무리\1. 작업 대기"
path_2 = r"C:\Users\jonkim\Desktop\리포트 마무리\1. 작업 대기\pd작업"
path_3 = r"C:\Users\jonkim\Desktop\리포트 마무리\GSA용"

f = []
f_name = []

# 작업 폴더에 있는 파일의 filenames 만 리스트에 저장 
for (dirpath, dirnames, filenames) in walk(path_1):
    f.extend(filenames)
    for i in filenames:
        temp_i = i.split("HDR")
        if temp_i[0] in f_name:
            continue
        else:
            f_name.append(temp_i[0])
    break # 이 break가 꼭 필요할까?

# 같은 이름의 파일들을 묶어주기
for j in f_name:
    group_f = []
    for q in f:
        temp_q = q.split("HDR")
        if j == temp_q[0]:
            group_f.append(q)
        else:
            continue

    excel = pd.DataFrame()
    for file_name in group_f:
        df = pd.read_excel(path_1 + "\\" + file_name)

        df.dropna(inplace=True)  # 빈 셀이 있는 행을 삭제
        # df.drop_duplicates(inplace=True)  # 중복 항목 제거

        # 제거할 도메인 리스트
        drop_list = ["Site URL",
                     #"simplesite.com",
                     #"edublogs.org",
                     #"penzu.com",
                     #"storeboard.com",
                     #"theglensecret.com",
                     #"theburnward.com",
                     #"fotosdefrases.com",
                     #"zenwriting.net",
                     #"onfeetnation.com",
                     #"writeablog.net",
                     #"cavandoragh.org",
                     #"evernote.com",
                     "postheaven.net",
                     #"bloggersdelight.dk"
                     "doodlekit.com"
                     ]

        # for문을 통해 도메인이 있는 행을 제거
        for h in drop_list:
            idx = df[df["Site URL"] == h].index     # Site URL 중에 j 값이 저장됨
            df.drop(idx, inplace=True)  # 해당 인덱스를 제거함

        # 필요없는 열 삭제
        df.drop(["Site URL", "D.A.", "Status", "Username", "Password", "E-Mail"],
                axis="columns", inplace=True)
        # 필요없는 행 삭제 (아래 도메인 제거 코드에서 작업 가능 함)
        # spac = df[df['Task ID'].str.contains('Task ID')].index
        # df.drop(spac, inplace=True)


        # 인덱스 재정렬
        # df.reset_index(drop=True, inplace=True)

        df.columns = ["Tier #", "Backlink List URL"]

        # 최종 정리후 저장
        excel = excel.append(df, ignore_index=True)

    # 데이터를 내림차순으로 정렬 (작업 속도가 상승하는지 체크 필요)
    # excel = excel.sort_values(by=["Task ID"], ascending=True)

    # 저장 전 Tasks 순서대로 정렬하기
    excel_2 = pd.DataFrame()
    text_GSA = pd.DataFrame()
    
    # 만약 Tasks 에 항목이 빠지면 진행 불가, 없으면 넘어가는 코드 작성 필요
    Tasks = ["Web 2.0 Blogs", "Social Bookmarking",
             "Authority Links", "Edu", "URL Shortener"]

    for t in Tasks:
        temp_task = excel[excel["Tier #"].str.contains(t)]
        excel_2 = excel_2.append(temp_task, ignore_index=True)
        if t == "Web 2.0 Blogs":
            text_GSA = text_GSA.append(temp_task, ignore_index=True)

    # 최종 데이터의 총 row 수를 저장
    las_row = len(excel_2)

    # 엑셀로 정리하기 전에 정렬하기, 이후 꾸미기까지
    excel_2.to_excel(f"{path_2}\\{j}{las_row}.xlsx", index=False)

    # GSA 필요없는 열 삭제
    text_GSA.drop(["Tier #"], axis="columns", inplace=True)
    # 인댁스 줄 삭제
    text_GSA.to_csv(f"{path_3}\\{j}.txt", index=False, header = None)
    

#---------------------------------------------------------


    wb = openpyxl.load_workbook(f"{path_2}\\{j}{las_row}.xlsx")
    ws = wb.active
   
    # 1차 가공 - 셀 정렬, 크기 조절, 폰트 변경

    for row in ws.rows:
        for cell in row:
            # 셀 폰트 변경
            cell.font = Font(name="Calibri", size=11)
            
    # column 크기 조절
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 130

    # 해더 색상 입히기
    for row in ws["A1:B1"]:
        for cell in row:
            cell.font = Font(name="Calibri", size=11, color=colors.WHITE, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="bottom")
            cell.fill = PatternFill(start_color="366092", fill_type="solid")


    # Task ID 이름들 제거하기
    column_max = ws.max_column
    row_max = ws.max_row
    Tasks2 = ["Web 2.0 Blogs ", "Social Bookmarking ",
             "Authority Links ", "Edu ", "URL Shortener "]

    for t2 in Tasks2:
        #열마다 행을 for loop문 진행
        for col_num in range(1, column_max+1):
            for row_num in range(1, row_max+1):
                
                #tempstr : cell값이 문자열이 아닌 경우를 감안하여 str로 바꿔줌
                tempstr = str(ws.cell(row = row_num, column = col_num).value)
                #문자열 함수 replace 사용(,를 빈칸으로)
                data = tempstr.replace(t2,"")
                #빈 셀의 경우 None이라는 문자열 타입이므로 제외하고 데이터 입력
                if data != "None":
                    ws.cell(row = row_num, column = col_num).value = data 
        
    
    # 저장하기
    wb.save(f"{path_2}\\{j}{las_row}.xlsx")

