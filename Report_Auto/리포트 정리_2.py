from asyncio import futures
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, colors, Font, Alignment
import openpyxl
import os
from os import walk

path_1 = r"C:\Users\kangg\Desktop\리포트 마무리\작업 대기\pd작업"
path_2 = r"C:\Users\kangg\Desktop\리포트 마무리"

f = []

# 작업 폴더에 있는 파일의 filenames 만 리스트에 저장 
for (dirpath, dirnames, filenames) in walk(path_1):
    f.extend(filenames)
    break # 이 break가 꼭 필요할까?

for full_name in f:
    wb = openpyxl.load_workbook(path_1 + "\\" + full_name)
    ws = wb.active

    # 1차 가공 - 셀 정렬, 크기 조절, 폰트 변경

    # 필요 없는 1번째 row 삭제
    ws.delete_rows(1)

    for row in ws.rows:
        for cell in row:
            # 셀 폰트 변경
            cell.font = Font(name="나눔고딕", size=10)
            
            # D.A column 중앙 정렬
            if cell.column == 3:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # column 크기 조절
    ws.column_dimensions["A"].width = 27.1
    ws.column_dimensions["B"].width = 27.1
    ws.column_dimensions["C"].width = 4.5
    ws.column_dimensions["D"].width = 50


    # 2차 가공 - 라벨 삽입하기

    # A열에서 특정 값을 가진 첫번 째 셀 row 좌표 찾기
    label_list = ["Web 2.0 Blogs", "Social Bookmarking", "Authority Links", "Edu", "URL Shortener"]

    for label in label_list:
        temp_x = ()
        for row in ws.iter_rows():
            temp_i = row[0].value.split(" (")[0]
            if temp_i == label:
                temp_x = row[0].row
                break

        ws.insert_rows(temp_x)  # temp_x위에 빈줄 추가

        for i in ("A", "B", "C", "D"):
            ws[f"{i}{temp_x}"].font = Font(name="나눔고딕", size=8.5, color=colors.WHITE, bold=True)
            ws[f"{i}{temp_x}"].alignment = Alignment(horizontal="center", vertical="center")
            ws[f"{i}{temp_x}"].fill = PatternFill(start_color="366092", fill_type="solid")

        data = {1: label, 2: "Site URL", 3: "D.A.", 4: "Submitted URL"}
        for col, value in data.items():
            ws.cell(row=temp_x, column=col, value=value)
        
    # 분석표, 순위 기록, 타이틀 삽입
    ws.insert_rows(1, 9)  # 1번 행 부터 빈줄 9개 추가

    # 해더 만들기
    for i in (1, 2):
        ws.merge_cells(f"A{i}:D{i}")  # row 병합 하기
        ws[f"A{i}"].font = Font(name="나눔고딕", size=8.5, color=colors.WHITE, bold=True)
        ws[f"A{i}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"A{i}"].fill = PatternFill(start_color="366092", fill_type="solid")

    ws["A1"] = "Website: "

    temp_name = full_name.split(".xlsx")
    ws["A2"] = temp_name[0]

    # 로고 삽입
    ws.merge_cells("A3:D6") # 로고 자리 병합 하기
    ws.add_image(Image("C:/Users/kangg/Desktop/마케스터즈 잡다/마케스터즈 로고 MAX.bmp"), "A3")  # 이미지를 A3 위치에 삽입

    # 작업 전, 후 분석표 넣기
    ws.merge_cells("A7:B7") # 로고 자리 병합 하기
    ws.merge_cells("C7:D7") # 로고 자리 병합 하기

    ws["A7"] = "작업 전"
    ws["A7"].font = Font(name="나눔고딕", size=10)
    ws["A7"].alignment = Alignment(horizontal="center", vertical="center")

    ws["C7"] = "작업 후"
    ws["C7"].font = Font(name="나눔고딕", size=10)
    ws["C7"].alignment = Alignment(horizontal="center", vertical="center")

    # 분석표 템플릿 넣기 (삽입후 1픽셀씩 오른쪽, 아래 이동 하면 좋겠음)
    ws.merge_cells("A8:B8") # 분석표 자리 병합 하기
    ws.merge_cells("C8:D8") # 분석표 자리 병합 하기
    ws.row_dimensions[8].height = 320
    ws.add_image(Image("C:/Users/kangg/Desktop/마케스터즈 잡다/분석표 템플릿.bmp"), "A8")  # 이미지를 A8 위치에 삽입
    ws.add_image(Image("C:/Users/kangg/Desktop/마케스터즈 잡다/분석표 템플릿.bmp"), "C8")  # 이미지를 A8 위치에 삽입

    # 검색순위 넣기
    ws.merge_cells("A9:B9") # 검색순위 자리 병합 하기
    ws["A9"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("C9:D9") # 검색순위 자리 병합 하기
    ws["C9"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[9].height = 100

    # 마지막 줄에 푸터 삽입
    footer_x = ws.max_row+1
    ws.merge_cells(f"A{footer_x}:D{footer_x}")  # 병합 하기
    ws[f"A{footer_x}"] = "마케스터즈 | Marketsters | https://marketsters.com/"
    ws[f"A{footer_x}"].font = Font(name="나눔고딕", size=8.5, color=colors.WHITE, bold=True)
    ws[f"A{footer_x}"].alignment = Alignment(vertical="center")
    ws[f"A{footer_x}"].fill = PatternFill(start_color="366092", fill_type="solid")
    
    # 저장하기
    wb.save(path_2 + "\\" + full_name)
