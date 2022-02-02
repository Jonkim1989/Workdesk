from asyncio import futures
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, colors, Font, Alignment
import openpyxl
import os
from os import walk

path_1 = r"C:\Users\kangg\Desktop\리포트 마무리\장기 작업"
path_2 = r"C:\Users\kangg\Desktop\리포트 마무리\장기 업데이트"

f = []

# 작업 폴더에 있는 파일의 filenames 만 리스트에 저장 
for (dirpath, dirnames, filenames) in walk(path_1):
    f.extend(filenames)
    break # 이 break가 꼭 필요할까?

for full_name in f:
    wb = openpyxl.load_workbook(path_1 + "\\" + full_name)
    ws = wb.active

    # 백링크 부분 지우기
    # A열에서 특정 값을 가진 첫번 째 셀 row 좌표 찾기
    
    top_label = ()
    for row in ws.iter_rows():
        if row[0].value == "Web 2.0 Blogs":
            top_label = row[0].row
            break

    ws.delete_rows(top_label, ws.max_row - top_label)  # 백링크 처음부터 푸터 전까지 삭제 후 당김

    # 저장하기

    temp_name = full_name.split(".xlsx")
    wb.save(path_2 + "\\" + temp_name[0] + " - 주간 업데이트.xlsx")
